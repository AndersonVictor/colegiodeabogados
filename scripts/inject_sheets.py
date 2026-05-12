"""
inject_sheets.py
----------------
Agrega (o reemplaza) las 5 hojas de datos hardcodeados al Excel
"DASHBOARD INGE.xlsx" para que el dashboard los lea desde ahí.

Hojas que crea / actualiza:
  - Alerta_Sobrecarga
  - Pob_Proyeccion
  - Jueces_Pob_2026
  - Personal_Regimen
  - Personal_Cargos

Uso:
    python scripts/inject_sheets.py
"""

import os
import io
import pandas as pd
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, "data", "DASHBOARD INGE.xlsx")


# ── 1. Alerta_Sobrecarga ──────────────────────────────────────────────────────
df_alerta = pd.DataFrame(
    [
        ("SALA CIVIL DE HUANCAYO", 2380, 3049, 28.1),
        ("1° JUZGADO PENAL UNIPERSONAL DE HUANCAYO (PROC. INMEDIATOS)", 1100, 2069, 88.1),
        ("3° JUZGADO PENAL UNIPERSONAL DE HUANCAYO (PROC. INMEDIATOS)", 1100, 1512, 37.5),
        ("2° JUZGADO PENAL UNIPERSONAL DE HUANCAYO (PROC. COMUNES)", 550, 1354, 146.2),
        ("4° JUZGADO PENAL UNIPERSONAL DE HUANCAYO (PROC. COMUNES)", 550, 1348, 145.1),
        ("2° JUZGADO DE INVESTIGACIÓN PREPARATORIA DE HUANCAYO (PROC. COMUNES)", 810, 1316, 62.5),
        ("1° JUZGADO DE INVESTIGACIÓN PREPARATORIA DE HUANCAYO (PROC. COMUNES)", 810, 1255, 54.9),
        ("6° JUZGADO DE INVESTIGACIÓN PREPARATORIA DE HUANCAYO (PROC. COMUNES)", 810, 1254, 54.8),
        ("2° JUZGADO DE FAMILIA DE HUANCAYO", 850, 1152, 35.5),
        (
            "JUZGADO DE INVESTIGACIÓN PREPARATORIA DE CHUPACA (PROC. INMEDIATOS) (PROC. COMUNES)",
            810, 1068, 31.9,
        ),
        (
            "5° JUZGADO DE INVESTIGACIÓN PREPARATORIA SUPRAPROVINCIAL ESPECIALIZADO EN DELITO DE CORRUPCIÓN DE FUNCIONARIOS DE HUANCAYO",
            180, 329, 82.8,
        ),
        (
            "8° JUZGADO DE INVESTIGACIÓN PREPARATORIA SUPRAPROVINCIAL ESPECIALIZADO EN DELITO DE CORRUPCIÓN DE FUNCIONARIOS DE HUANCAYO",
            180, 310, 72.2,
        ),
        (
            "6° JUZGADO PENAL UNIPERSONAL SUPRAPROVINCIAL ESPECIALIZADO EN DELITOS DE CORRUPCIÓN DE FUNCIONARIOS HUANCAYO",
            90, 296, 228.9,
        ),
        (
            "5° JUZGADO PENAL UNIPERSONAL SUPRAPROVINCIAL ESPECIALIZADO EN DELITOS DE CORRUPCIÓN DE FUNCIONARIOS HUANCAYO",
            90, 267, 196.7,
        ),
    ],
    columns=["ORGANO_JURISDICCIONAL", "CARGA_MAXIMA", "CARGA_PROYECTADA", "PORC_SOBRECARGA"],
)

# ── 2. Pob_Proyeccion ─────────────────────────────────────────────────────────
df_pob_proy = pd.DataFrame(
    [
        ("Huancayo",       1.0,  630_005, 636_085, 642_224, 648_422, 654_680, 660_997),
        ("Concepción",    -0.3,   57_682,  57_531,  57_381,  57_231,  57_082,  56_933),
        ("Jauja",         -0.4,   85_233,  84_879,  84_527,  84_176,  83_826,  83_478),
        ("Junín (prov.)", -2.0,   19_472,  19_085,  18_706,  18_334,  17_970,  17_613),
        ("Tarma",         -1.4,   81_895,  80_711,  79_545,  78_396,  77_263,  76_147),
        ("Yauli",         -1.5,   35_505,  34_965,  34_434,  33_910,  33_395,  32_887),
        ("Chupaca",        0.5,   59_563,  59_871,  60_181,  60_492,  60_805,  61_120),
        ("Tayacaja",      -1.6,   75_599,  74_426,  73_272,  72_135,  71_016,  69_915),
    ],
    columns=["Provincia", "TasaCrec", "2026", "2027", "2028", "2029", "2030", "2031"],
)

# ── 3. Jueces_Pob_2026 ────────────────────────────────────────────────────────
df_jueces = pd.DataFrame(
    [
        ("A NIVEL NACIONAL (1)",            3776, 34_038_457, 11),
        ("DISTRITO JUDICIAL DE JUNÍN (2)",   119,  1_044_954, 11),
        ("Huancayo (3)",                      78,    630_005, 12),
        ("Concepción",                         4,     57_682,  4),
        ("Jauja",                              7,     85_233,  7),
        ("Junín",                              4,     19_472,  4),
        ("Tarma",                             12,     81_895, 12),
        ("Yauli",                              4,     35_505,  4),
        ("Chupaca",                            5,     59_563,  5),
        ("Tayacaja - Dpto. Huancavelica*",     5,     75_599,  5),
    ],
    columns=[
        "CATEGORIA",
        "CANTIDAD_JUECES",
        "PROY_POB_2026",
        "JUECES_POR_100K",
    ],
)

# ── 4. Personal_Regimen ───────────────────────────────────────────────────────
df_regimen = pd.DataFrame(
    {
        "Regimen":  ["LEY 29277", "728", "276", "CAS", "RECAS"],
        "Cantidad": [121,         482,   3,     331,   137],
    }
)

# ── 5. Personal_Cargos ────────────────────────────────────────────────────────
raw = """CARGO\tLEY 29277\t728\t276\tCAS\tRECAS\tTotal
ADMINISTRADOR\t\t4\t\t\t\t4
ADMINISTRADOR DE CORTE SUPERIOR\t\t1\t\t\t\t1
ADMINISTRADOR DE MODULO\t\t\t\t2\t\t2
ADMINISTRADOR I\t\t1\t\t\t\t1
ADMINISTRADOR MÓDULO DEL NUEVO CÓDIGO PROCESAL PENAL\t\t1\t\t\t\t1
AGENTE DE  SEGURIDAD\t\t\t\t7\t\t7
ANALISTA\t\t\t\t1\t1\t2
ANALISTA ADMINISTRATIVO/A\t\t\t\t2\t\t2
ANALISTA DE INFORMATICA\t\t\t\t1\t\t1
ANALISTA DE INFORMATICA - DESARROLLADOR\t\t\t\t1\t\t1
ANALISTA I\t\t1\t\t\t1\t2
ANALISTA II\t\t7\t\t\t\t7
APOYO ADMINISTRATIVO\t\t\t\t3\t1\t4
APOYO ADMINISTRATIVO EN EL ARCHIVO DESCENTRALIZADO DE TARMA Y OTRAS ACTIVIDADES ADMINISTRATIVAS\t\t\t\t\t1\t1
APOYO ADMINISTRATIVO EN EL AREA DE NOTIFICACIONES\t\t\t\t\t1\t1
APOYO ADMINISTRATIVO EN LAS AREAS DE NOTIFICACIONES\t\t\t\t\t3\t3
APOYO EN ACTIVIDADES ADMINISTRATIVAS\t\t\t\t1\t\t1
APOYO EN CONTRATACIONES CON EL ESTADO\t\t\t\t\t1\t1
APOYO EN EL DESARROLLO DE FUNCIONES DEL EQUIPO ITINERANTE\t\t\t\t\t3\t3
APOYO EN INFORMES PSICOLOGICOS\t\t\t\t\t1\t1
APOYO EN LA PREPARACION DE ALIMENTOS EN EL WAWA WASI INSTITUCIONAL\t\t\t\t\t1\t1
APOYO EN PERICIAS CONTABLES JUDICIALES\t\t\t\t1\t1\t2
APOYO JURISDICCIONAL\t\t\t\t\t2\t2
ASESOR DE CORTE\t\t1\t\t\t\t1
ASISTENTE ADMINISTRATIVO\t\t\t\t6\t1\t7
ASISTENTE ADMINISTRATIVO I\t\t14\t\t4\t\t18
ASISTENTE ADMINISTRATIVO I MODULO\t\t\t\t4\t\t4
ASISTENTE ADMINISTRATIVO II\t\t18\t\t1\t\t19
ASISTENTE DE ATENCION AL PUBLICO\t\t\t\t1\t\t1
ASISTENTE DE ATENCIÓN AL PÚBLICO\t\t\t\t4\t\t4
ASISTENTE DE COMUNICACIONES\t\t\t\t15\t2\t17
ASISTENTE DE CUSTODIA Y GRABACION\t\t\t\t2\t1\t3
ASISTENTE DE CUSTODIA Y GRABACIÓN\t\t\t\t4\t\t4
ASISTENTE DE INFORMÁTICA\t\t\t\t4\t\t4
ASISTENTE DE SISTEMAS\t\t2\t\t1\t\t3
ASISTENTE EN INFORMATICA\t\t\t\t\t1\t1
ASISTENTE EN SERVICIOS ADMINISTRATIVOS\t\t5\t\t1\t\t6
ASISTENTE EN SERVICIOS DE COMUNICACIONES\t\t13\t\t\t\t13
ASISTENTE INFORMÁTICO\t\t\t\t3\t\t3
ASISTENTE JUDICIAL\t\t55\t\t20\t2\t77
ASISTENTE JUDICIAL - PENAL\t\t\t\t6\t\t6
ASISTENTE JUDICIAL - PROTECCIÓN\t\t\t\t11\t\t11
ASISTENTE JURISDICCIONAL\t\t13\t\t20\t2\t35
ASISTENTE JURISDICCIONAL DE CUSTODIA DE EXPEDIENTES\t\t3\t\t\t\t3
ASISTENTE JURISDICCIONAL DE JUZGADO\t\t\t\t22\t12\t34
ASISTENTE JURISDICCIONAL ITINERANTE\t\t\t\t\t2\t2
ASISTENTE LEGAL\t\t\t\t9\t\t9
ASISTENTE LEGAL DE JUZGADO\t\t\t\t1\t\t1
ASISTENTE SOCIAL\t\t\t\t2\t1\t3
AUXILIAR ADMINISTRATIVO\t\t\t\t3\t\t3
AUXILIAR ADMINISTRATIVO I\t\t44\t\t\t\t44
AUXILIAR ADMINISTRATIVO II\t\t6\t\t\t\t6
AUXILIAR EN VIGILANCIA Y CONTROL\t\t\t\t6\t\t6
AUXILIAR JUDICIAL\t\t37\t\t10\t\t47
AUXILIAR JUDICIAL - PROTECCIÓN\t\t\t\t5\t\t5
AUXILIAR JUDICIAL I\t\t\t1\t\t\t1
AUXILIAR JURISDICCIONAL ITINERANTE\t\t\t\t\t1\t1
AUXILIAR LEGAL\t\t\t\t1\t\t1
CAJERO\t\t1\t\t\t\t1
CHOFER\t\t\t\t2\t1\t3
CHOFER I\t\t3\t\t\t\t3
CONDUCCION DE VEHICULO\t\t\t\t1\t4\t5
CONDUCTOR DE VEHÍCULO\t\t\t\t1\t\t1
COORDINADOR(CSJ_UE)\t\t16\t\t\t\t16
COORDINADOR/A DE CAUSA/ AUDIENCIA\t\t\t\t3\t\t3
COORDINADOR/A I\t\t\t\t2\t\t2
EDUCADOR (A) DE APOYO AL EQUIPO MULTIDISCIPLINARIO\t\t\t\t\t1\t1
ESPECIALISTA EN CONTABILIDAD GUBERNAMENTAL\t\t\t\t\t1\t1
ESPECIALISTA JUDICIAL DE AUDIENCIA\t\t19\t\t16\t11\t46
ESPECIALISTA JUDICIAL DE JUZGADO\t\t30\t\t46\t24\t100
ESPECIALISTA JUDICIAL DE SALA\t\t4\t\t1\t\t5
ESPECIALISTA LEGAL\t\t\t\t11\t\t11
ESPECIALISTA LEGAL DE AUDIENCIA\t\t\t\t2\t\t2
ESPECIALISTA LEGAL DE JUZGADO\t\t\t\t3\t\t3
ESPECIALISTA LEGAL DE SALA\t\t\t\t3\t\t3
JEFE DE UNIDAD\t\t3\t\t\t\t3
JEFE/A DE UNIDAD\t\t\t\t2\t\t2
JUEZ DE PAZ LETRADO\t20\t\t\t\t\t20
JUEZ ESPECIALIZADO\t72\t\t\t\t\t72
JUEZ ESPECIALIZADO 7U-4\t6\t\t\t\t\t6
JUEZ SUPERIOR\t22\t\t\t\t\t22
MADRE CUIDADORA DEL WAWA WASI INSTITUCIONAL\t\t\t\t\t2\t2
MEDICO\t\t\t\t\t1\t1
NOTIFICADOR\t\t\t\t6\t3\t9
PERITO JUDICIAL\t\t2\t\t2\t\t4
PERSONAL  DE SEGURIDAD\t\t\t\t6\t\t6
PRESIDENTE DE CORTE SUPERIOR\t1\t\t\t\t\t1
PROFESIONAL LEGAL\t\t\t\t2\t\t2
PROFESIONAL LEGAL DE JUZGADO\t\t\t\t1\t\t1
PSICOLOGO\t\t5\t\t5\t4\t14
RELATOR I\t\t4\t\t\t\t4
RESGUARDO, CUSTODIA Y VIGILANCIA\t\t\t\t6\t36\t42
REVISOR\t\t1\t\t\t\t1
SECRETARIO DE SALA\t\t4\t\t\t\t4
SECRETARIO III\t\t1\t\t\t\t1
SECRETARIO JUDICIAL\t\t127\t\t4\t4\t135
SECRETARIO JUDICIAL I\t\t\t1\t\t\t1
SECRETARIO/A DE JUZGADO - PROTECCIÓN\t\t\t\t7\t\t7
SECRETARIO/A JUDICIAL\t\t\t\t11\t\t11
SEGURIDAD\t\t\t\t1\t\t1
TECNICO ADMINISTRATIVO I\t\t1\t\t\t\t1
TECNICO ADMINISTRATIVO II\t\t1\t\t1\t\t2
TECNICO JUDICIAL\t\t26\t\t\t\t26
TECNICO JUDICIAL I\t\t\t1\t\t\t1
TRABAJADOR SOCIAL\t\t8\t\t\t\t8
TRABAJADORA SOCIAL\t\t\t\t1\t3\t4"""

df_cargos = pd.read_csv(io.StringIO(raw), sep="\t").fillna(0)
for col in ["LEY 29277", "728", "276", "CAS", "RECAS", "Total"]:
    df_cargos[col] = df_cargos[col].astype(int)


# ── Escribir en Excel ─────────────────────────────────────────────────────────
SHEET_MAP = {
    "Alerta_Sobrecarga": df_alerta,
    "Pob_Proyeccion":    df_pob_proy,
    "Jueces_Pob_2026":   df_jueces,
    "Personal_Regimen":  df_regimen,
    "Personal_Cargos":   df_cargos,
}

wb = load_workbook(EXCEL_PATH)

for sheet_name, df in SHEET_MAP.items():
    # Eliminar hoja si ya existe para reescribirla
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    # Encabezados
    ws.append(list(df.columns))

    # Filas
    for row in df.itertuples(index=False):
        ws.append(list(row))

    print(f"  Hoja '{sheet_name}': {len(df)} filas escritas.")

wb.save(EXCEL_PATH)
print(f"\nExcel guardado en: {EXCEL_PATH}")
print("Hojas disponibles:", wb.sheetnames)
